"""Export leads to various formats"""
import csv
import io
from typing import List
from app.core.models import Lead


class ExportService:
    """Service for exporting leads to different formats"""
    
    @staticmethod
    def to_csv(leads: List[Lead]) -> str:
        """Export leads to CSV format"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            "Name",
            "Niche",
            "Website",
            "Email",
            "Phone",
            "Address",
            "City",
            "Country",
            "Source",
            "Quality Score",
            "Quality Label",
            "Social Links",
        ])
        
        # Write data
        for lead in leads:
            writer.writerow([
                lead.name or "",
                lead.niche or "",
                lead.website or "",
                "; ".join(lead.emails or []),
                "; ".join(lead.phones or []),
                lead.address or "",
                lead.city or "",
                lead.country or "",
                lead.source or "",
                str(lead.quality_score) if hasattr(lead, 'quality_score') and lead.quality_score else "",
                lead.quality_label if hasattr(lead, 'quality_label') else "",
                ", ".join([f"{k}: {v}" for k, v in (lead.social_links or {}).items()]),
            ])
        
        return output.getvalue()
    
    @staticmethod
    def to_excel(leads: List[Lead]) -> bytes:
        """Export leads to Excel format (requires openpyxl)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            
            # Header row with styling
            headers = [
                "Name", "Niche", "Website", "Email", "Phone", "Address",
                "City", "Country", "Source", "Quality Score", "Quality Label", "Social Links"
            ]
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Write data
            for lead in leads:
                ws.append([
                    lead.name or "",
                    lead.niche or "",
                    lead.website or "",
                    "; ".join(lead.emails or []),
                    "; ".join(lead.phones or []),
                    lead.address or "",
                    lead.city or "",
                    lead.country or "",
                    lead.source or "",
                    lead.quality_score if hasattr(lead, 'quality_score') and lead.quality_score else "",
                    lead.quality_label if hasattr(lead, 'quality_label') else "",
                    ", ".join([f"{k}: {v}" for k, v in (lead.social_links or {}).items()]),
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
            
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
